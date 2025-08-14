def generar_excel_devolucion(fecha, proveedor, df, nombre_reporte="Registro_Devolucion_Proveedores"):
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_reporte.replace("_", " ")

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.cell(row=1, column=1, value="ANEXO B: Formato de Registro de Devolución a Proveedores").font = bold
    ws.cell(row=1, column=1).alignment = center

    ws.cell(row=2, column=1, value="PROGRAMA DE CONTROL DE PROVEEDORES").font = bold
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws.cell(row=2, column=1).alignment = center

    ws.cell(row=3, column=1, value="FORMATO DE REGISTRO DE DEVOLUCIÓN A PROVEEDORES").font = bold
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
    ws.cell(row=3, column=1).alignment = center

    # Datos generales
    ws.cell(row=4, column=1, value="Fecha de Devolución:")
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
    ws.cell(row=4, column=3, value=str(fecha))
    ws.cell(row=4, column=4, value="Proveedor:")
    ws.cell(row=4, column=5, value=proveedor)

    # Encabezados
    headers = [
        "Producto(s) Devuelto(s)", "Cantidad", "Lote", "Causal del Rechazo (Marcar con X)"
    ]
    for idx, h in enumerate(headers, 1):
        ws.cell(row=6, column=idx, value=h).font = bold
        ws.cell(row=6, column=idx).alignment = center
        ws.cell(row=6, column=idx).border = border

    # Llenar datos de la tabla
    for i, row in enumerate(df.values.tolist()):
        for j, value in enumerate(row):
            cell = ws.cell(row=7 + i, column=1 + j, value=value)
            cell.alignment = center
            cell.border = border

    # Espacio para firmas
    row_firma = 8 + len(df)
    ws.cell(row=row_firma, column=1, value="Responsable de la Devolución:")
    ws.merge_cells(start_row=row_firma, start_column=2, end_row=row_firma, end_column=3)
    ws.cell(row=row_firma, column=2, value="________________________")
    ws.cell(row=row_firma, column=4, value="Firma:")
    ws.merge_cells(start_row=row_firma, start_column=5, end_row=row_firma, end_column=5)
    ws.cell(row=row_firma, column=5, value="______________")

    ws.cell(row=row_firma+1, column=1, value="Nombre del Conductor/Representante del Proveedor:")
    ws.merge_cells(start_row=row_firma+1, start_column=2, end_row=row_firma+1, end_column=3)
    ws.cell(row=row_firma+1, column=2, value="________________________")
    ws.cell(row=row_firma+1, column=4, value="Firma:")
    ws.merge_cells(start_row=row_firma+1, start_column=5, end_row=row_firma+1, end_column=5)
    ws.cell(row=row_firma+1, column=5, value="______________")

    # Ajustar ancho de columnas automáticamente
    from openpyxl.utils import get_column_letter
    for idx, col in enumerate(ws.columns, 1):
        max_length = 0
        col_letter = get_column_letter(idx)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max(12, max_length + 2)

    # Guardar archivo temporal
    fecha_str = str(fecha)
    nombre_archivo = f"{nombre_reporte}_{fecha_str}.xlsx"
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx", prefix=nombre_archivo.replace(" ", "_"))
    wb.save(tmp.name)
    return tmp.name

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from tempfile import NamedTemporaryFile

def generar_excel(nombre_establecimiento, fecha, df, responsable, supervisor, revision, nombre_reporte="Verificacion_BPM"):
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_reporte.replace("_", " ")

    # Estilos
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal="center", vertical="center")

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.cell(row=1, column=1, value="FORMATO DE INSPECCIÓN Y CONTROL DE RECEPCIÓN DE MATERIAS PRIMAS").font = bold
    ws.cell(row=1, column=1).alignment = center

    # Fila de datos generales
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    ws.cell(row=2, column=1, value="Fecha:")
    ws.cell(row=2, column=3, value=str(fecha))
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=5)
    ws.cell(row=2, column=4, value="Proveedor:")
    ws.cell(row=2, column=6, value=nombre_establecimiento)
    ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=7)
    ws.cell(row=2, column=7, value=f"Factura N°: {supervisor if supervisor else ''}")


    # Encabezados de la tabla (dos filas)
    ws.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)
    ws.cell(row=4, column=1, value="Producto").font = bold
    ws.cell(row=4, column=1).alignment = center
    ws.cell(row=4, column=1).border = border

    ws.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)
    ws.cell(row=4, column=2, value="Lote").font = bold
    ws.cell(row=4, column=2).alignment = center
    ws.cell(row=4, column=2).border = border

    ws.merge_cells(start_row=4, start_column=3, end_row=5, end_column=3)
    ws.cell(row=4, column=3, value="Fecha Vencimiento").font = bold
    ws.cell(row=4, column=3).alignment = center
    ws.cell(row=4, column=3).border = border

    ws.merge_cells(start_row=4, start_column=4, end_row=5, end_column=4)
    ws.cell(row=4, column=4, value="Temp. (°C)").font = bold
    ws.cell(row=4, column=4).alignment = center
    ws.cell(row=4, column=4).border = border

    ws.merge_cells(start_row=4, start_column=5, end_row=5, end_column=5)
    ws.cell(row=4, column=5, value="Características Organolépticas (Color, Olor, Textura)").font = bold
    ws.cell(row=4, column=5).alignment = center
    ws.cell(row=4, column=5).border = border

    ws.merge_cells(start_row=4, start_column=6, end_row=5, end_column=6)
    ws.cell(row=4, column=6, value="Empaque").font = bold
    ws.cell(row=4, column=6).alignment = center
    ws.cell(row=4, column=6).border = border

    ws.merge_cells(start_row=4, start_column=7, end_row=5, end_column=7)
    ws.cell(row=4, column=7, value="Observaciones").font = bold
    ws.cell(row=4, column=7).alignment = center
    ws.cell(row=4, column=7).border = border

    # Llenar datos de la tabla
    start_row = 6
    for i, row in enumerate(df.values.tolist()):
        for j, value in enumerate(row):
            cell = ws.cell(row=start_row + i, column=1 + j, value=value)
            cell.alignment = center
            cell.border = border

    # Bordes para filas vacías si hay menos de 15 productos (opcional, puedes quitarlo si quieres solo filas dinámicas)
    # for i in range(len(df), 15):
    #     for j in range(6):
    #         cell = ws.cell(row=start_row + i, column=1 + j)
    #         cell.border = border

    # Condiciones del vehículo
    row_cond = start_row + max(len(df), 10) + 2
    ws.merge_cells(start_row=row_cond, start_column=1, end_row=row_cond, end_column=7)
    ws.cell(row=row_cond, column=1, value=f"Condiciones del Vehículo de Transporte: Limpio: {responsable} | Temperatura del furgón (si aplica): {revision} °C").font = bold

    # Recibido por y firma
    row_firma = row_cond + 2
    ws.cell(row=row_firma, column=1, value="Recibido por:")
    ws.merge_cells(start_row=row_firma, start_column=2, end_row=row_firma, end_column=4)
    ws.cell(row=row_firma, column=2, value="________________________")
    ws.cell(row=row_firma, column=5, value="Firma:")
    ws.merge_cells(start_row=row_firma, start_column=6, end_row=row_firma, end_column=7)
    ws.cell(row=row_firma, column=6, value="______________")

    # Ajustar ancho de columnas automáticamente
    from openpyxl.utils import get_column_letter
    for idx, col in enumerate(ws.columns, 1):
        max_length = 0
        col_letter = get_column_letter(idx)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max(12, max_length + 2)

    # Guardar archivo temporal
    fecha_str = str(fecha)
    nombre_archivo = f"{nombre_reporte}_{fecha_str}.xlsx"
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx", prefix=nombre_archivo.replace(" ", "_"))
    wb.save(tmp.name)
    return tmp.name